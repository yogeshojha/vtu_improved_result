#Author: Yogesh Ojha
#FEB 12 2015

import requests
from lxml import html
import openpyxl
import pprint
from collections import OrderedDict
import sys

BASE_URL = 'http://results.vtu.ac.in'

def get_result(usn):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.10; rv:39.0) Gecko/20100101 Firefox/39.0',
        'X-Requested-With': 'XMLHttpRequest',
        'Host': 'results.vtu.ac.in',
        'Referer': 'http://results.vtu.ac.in/'
    }

    payload = {
        'rid': usn,
        'submit': 'SUBMIT'
    }

    # xpath selector for subject name
    sub_xpath = '/html/body/table/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[2]/td[1]/table[2]/tbody/tr/td/table/tbody/tr[2]/td/table[2]/tr[{}]/td[{}]/i/text()'

    # xpath selector for subject external marks, internal marks and total marks
    sub_xpath2 = '/html/body/table/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[2]/td[1]/table[2]/tbody/tr/td/table/tbody/tr[2]/td/table[2]/tr[{}]/td[{}]/text()'

    # xpath selector for subject result
    sub_xpath3 = '/html/body/table/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[2]/td[1]/table[2]/tbody/tr/td/table/tbody/tr[2]/td/table[2]/tr[{}]/td[{}]/b/text()'

    response = requests.post(BASE_URL + '/vitavi.php', payload, headers=headers)
    tree = html.fromstring(response.content)

    # student details
    student_name_usn = tree.xpath('/html/body/table/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[2]/td[1]/table[2]/tbody/tr/td/table/tbody/tr[2]/td/b/text()')
    total_marks = tree.xpath('/html/body/table/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[2]/td[1]/table[2]/tbody/tr/td/table/tbody/tr[2]/td/table[3]/tr/td[4]/text()')
    semester = tree.xpath('/html/body/table/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[2]/td[1]/table[2]/tbody/tr/td/table/tbody/tr[2]/td/table[1]/tr/td[2]/b/text()')
    result = tree.xpath('/html/body/table/tbody/tr[3]/td[2]/table/tbody/tr[3]/td/table/tbody/tr[2]/td[1]/table[2]/tbody/tr/td/table/tbody/tr[2]/td/table[1]/tr/td[4]/b/text()')

    # subject details
    sub1 = tree.xpath(sub_xpath.format('2', '1'))
    sub1_external = tree.xpath(sub_xpath2.format('2', '2'))
    sub1_internal = tree.xpath(sub_xpath2.format('2', '3'))
    sub1_total = tree.xpath(sub_xpath2.format('2', '4'))
    sub1_result = tree.xpath(sub_xpath3.format('2', '5'))

    sub2 = tree.xpath(sub_xpath.format('3', '1'))
    sub2_external = tree.xpath(sub_xpath2.format('3', '2'))
    sub2_internal = tree.xpath(sub_xpath2.format('3', '3'))
    sub2_total = tree.xpath(sub_xpath2.format('3', '4'))
    sub2_result = tree.xpath(sub_xpath3.format('3', '5'))

    sub3 = tree.xpath(sub_xpath.format('4', '1'))
    sub3_external = tree.xpath(sub_xpath2.format('4', '2'))
    sub3_internal = tree.xpath(sub_xpath2.format('4', '3'))
    sub3_total = tree.xpath(sub_xpath2.format('4', '4'))
    sub3_result = tree.xpath(sub_xpath3.format('4', '5'))

    sub4 = tree.xpath(sub_xpath.format('5', '1'))
    sub4_external = tree.xpath(sub_xpath2.format('5', '2'))
    sub4_internal = tree.xpath(sub_xpath2.format('5', '3'))
    sub4_total = tree.xpath(sub_xpath2.format('5', '4'))
    sub4_result = tree.xpath(sub_xpath3.format('5', '5'))

    sub5 = tree.xpath(sub_xpath.format('6', '1'))
    sub5_external = tree.xpath(sub_xpath2.format('6', '2'))
    sub5_internal = tree.xpath(sub_xpath2.format('6', '3'))
    sub5_total = tree.xpath(sub_xpath2.format('6', '4'))
    sub5_result = tree.xpath(sub_xpath3.format('6', '5'))

    sub6 = tree.xpath(sub_xpath.format('7', '1'))
    sub6_external = tree.xpath(sub_xpath2.format('7', '2'))
    sub6_internal = tree.xpath(sub_xpath2.format('7', '3'))
    sub6_total = tree.xpath(sub_xpath2.format('7', '4'))
    sub6_result = tree.xpath(sub_xpath3.format('7', '5'))

    sub7 = tree.xpath(sub_xpath.format('8', '1'))
    sub7_external = tree.xpath(sub_xpath2.format('8', '2'))
    sub7_internal = tree.xpath(sub_xpath2.format('8', '3'))
    sub7_total = tree.xpath(sub_xpath2.format('8', '4'))
    sub7_result = tree.xpath(sub_xpath3.format('8', '5'))

    sub8 = tree.xpath(sub_xpath.format('9', '1'))
    sub8_external = tree.xpath(sub_xpath2.format('9', '2'))
    sub8_internal = tree.xpath(sub_xpath2.format('9', '3'))
    sub8_total = tree.xpath(sub_xpath2.format('9', '4'))
    sub8_result = tree.xpath(sub_xpath3.format('9', '5'))

    try:
        student_name_usn = student_name_usn[0].split('(')
        result = result[0].split()
        total_marks = total_marks[0].strip().replace(' ','')
        if len(result) == 3:
            result = result[1] + ' ' + result[2]
        else:
            result = result[1]

        student = OrderedDict([
            ('name', student_name_usn[0].strip()),
            ('usn', student_name_usn[1].strip().replace(')', '')),
            ('semester', semester[0]),
            ('result', result),
            ('total_marks', total_marks),
            ('subject_marks', [
                    OrderedDict([
                        ('subject', sub1[0]),
                        ('external', sub1_external[0]),
                        ('internal', sub1_internal[0]),
                        ('total', sub1_total[0]),
                        ('result', sub1_result[0])
                    ]),
                    OrderedDict([
                        ('subject', sub2[0]),
                        ('external', sub2_external[0]),
                        ('internal', sub2_internal[0]),
                        ('total', sub2_total[0]),
                        ('result', sub2_result[0])
                    ]),
                    OrderedDict([
                        ('subject', sub3[0]),
                        ('external', sub3_external[0]),
                        ('internal', sub3_internal[0]),
                        ('total', sub3_total[0]),
                        ('result', sub3_result[0])
                    ]),
                    OrderedDict([
                        ('subject', sub4[0]),
                        ('external', sub4_external[0]),
                        ('internal', sub4_internal[0]),
                        ('total', sub4_total[0]),
                        ('result', sub4_result[0])
                    ]),
                    OrderedDict([
                        ('subject', sub5[0]),
                        ('external', sub5_external[0]),
                        ('internal', sub5_internal[0]),
                        ('total', sub5_total[0]),
                        ('result', sub5_result[0])
                    ]),
                    OrderedDict([
                        ('subject', sub6[0]),
                        ('external', sub6_external[0]),
                        ('internal', sub6_internal[0]),
                        ('total', sub6_total[0]),
                        ('result', sub6_result[0])
                    ]),
                    OrderedDict([
                        ('subject', sub7[0]),
                        ('external', sub7_external[0]),
                        ('internal', sub7_internal[0]),
                        ('total', sub7_total[0]),
                        ('result', sub7_result[0])
                    ]),
                    OrderedDict([
                        ('subject', sub8[0]),
                        ('external', sub8_external[0]),
                        ('internal', sub8_internal[0]),
                        ('total', sub8_total[0]),
                        ('result', sub8_result[0])
                    ]),
                ])
            ])

    except IndexError as e:
       #print("USN doesn't exist\n")
        student = None

    return student

def create_excel_file(filename):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'] = 'USN'
    sheet['B1'] = 'Student'
    sheet['C1'] = 'SEMESTER'
    sheet['D1'] = 'RESULT'
    sheet['E1'] = 'TOTAL MARKS'

    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15

    wb.save(filename + '.xlsx')

def write_to_excel_file(students, filename):
    wb = openpyxl.load_workbook(filename + '.xlsx')
    sheet = wb.get_sheet_by_name('Sheet')
    
    row_num = 2
    for student in students:
        sheet.cell(row=row_num, column=1).value = student['usn']
        sheet.cell(row=row_num, column=2).value = student['name']
        sheet.cell(row=row_num, column=3).value = student['semester']
        sheet.cell(row=row_num, column=4).value = student['result']
        sheet.cell(row=row_num, column=5).value = student['total_marks']
        row_num = row_num + 1

    print("Exporting to excel file: %s.xlsx" % filename)
    wb.save(filename + '.xlsx')

def main():
    print "Exports the results of vtu students into Excel format \n"
    print "Please note the format of USN is College_codeYearBranchUsn\nExample: 1EP14CS001\nwhere 1EP is the college code\n14 is the batch year\nCS is the branch\n001 is the usn"
    usn = raw_input("Enter the starting USN\nExample: 1EP14CS001\n")
    to = int(raw_input("Enter final USN\nEnter the last 3 digits of final usn number.\nExample: 100.\nThis will be treated as 1EP14CS100\n"))
    usn_part = usn[0:7]
    frm = int(usn[8:10])
    filename = usn[1:3]+'_'+usn[5:7]+'_'+'results'
    create_excel_file(filename)
    students = []
    for i in range(int(frm), int(to) + 1):
        student = get_result(usn_part + str(i).zfill(3))
        if student is not None:
            print("Getting result of %s (%s)\n" % (student['name'], student['usn']))
            students.append(student)
    write_to_excel_file(students, filename)

if __name__ == '__main__':
    main()
